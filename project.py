 #Import required library
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
# Specify the path to your Excel file
filepath = "/Users/ehf/Documents/GitHub/KC-final-project/essa.xlsx"

# Load the workbook
wb = load_workbook(filepath)
Full_Sheet = wb['Input']

last_row = Full_Sheet.max_row
last_column = Full_Sheet.max_column

# Create destination workbook
dest_workbook = openpyxl.Workbook()

# Delete existing sheets
for sheet in dest_workbook.sheetnames:
    del dest_workbook[sheet]

#---------------------------- Letter Grades Sheet ---------------------------

# Create new sheet with name "Letter Grades"
dest_sheet = dest_workbook.create_sheet("Letter Grades")

for row in Full_Sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=last_column):
    for cell in row:
        dest_sheet[cell.coordinate] = cell.value

dest_sheet.cell(row=1, column=last_column + 1, value="Total")
dest_sheet.cell(row=1, column=last_column + 2, value="Letter Grade")
dest_sheet.cell(row=1, column=last_column + 3, value="Pass / Fail")

# Perform calculations and populate column 18
for row_num in range(4, last_row + 1):  # Loop through each row from row 4 to last_row
    total_assignment = sum([dest_sheet.cell(row=row_num, column=col_num).value for col_num in range(3, 8)])  # Sum from column 3 to column 7
    calculated_value_assignment = (total_assignment / .5) * dest_sheet.cell(row=2, column=3).value  # Perform the calculation

    total_quiz = sum([dest_sheet.cell(row=row_num, column=col_num).value for col_num in range(8, 13)])  # Sum from column 8 to column 12
    calculated_value_quiz = (total_quiz / .5) * dest_sheet.cell(row=2, column=8).value  # Perform the calculation

    total_midterm = sum([dest_sheet.cell(row=row_num, column=col_num).value for col_num in range(13, 15)])  # Sum from column 13 to column 14
    calculated_value_midterm = (total_midterm / 2) * dest_sheet.cell(row=2, column=13).value  # Perform the calculation

    calculated_value_presenation = dest_sheet.cell(row=row_num, column=15).value * dest_sheet.cell(row=2, column=15).value # Perform the calculation

    calculated_value_project = dest_sheet.cell(row=row_num, column=16).value * dest_sheet.cell(row=2, column=16).value  # Perform the calculation

    calculated_value_final_exam = dest_sheet.cell(row=row_num, column=17).value * dest_sheet.cell(row=2, column=17).value  # Perform the calculation

    dest_sheet.cell(row=row_num, column=18, value=calculated_value_assignment+calculated_value_quiz+calculated_value_midterm+calculated_value_presenation+calculated_value_project+calculated_value_final_exam)  # Place the result in column 18

# Apply formula from row 4 to last row
for row in range(4, last_row + 1):
    cell_to_check = f'S{row}'
    cell_to_write = f'T{row}'
    dest_sheet[cell_to_write].value = f'=IF({cell_to_check}="F", "Fail", "Pass")'

# Assign letter grades based on total score (from column 18)
for row_num in range(4, last_row + 1):
    total_score = dest_sheet.cell(row=row_num, column=18).value  # Retrieve the total score from column 18

    if total_score is not None:
        if total_score >= 90:
            grade = 'A'
        elif total_score >= 85:
            grade = 'A-'
        elif total_score >= 80:
            grade = 'B+'
        elif total_score >= 75:
            grade = 'B'
        elif total_score >= 70:
            grade = 'B-'
        elif total_score >= 65:
            grade = 'C+'
        elif total_score >= 60:
            grade = 'C'
        elif total_score >= 55:
            grade = 'C-'
        elif total_score >= 50:
            grade = 'D'
        else:
            grade = 'F'

        dest_sheet.cell(row=row_num, column=19, value=grade)  # Store the letter grade in column 19

# Create a border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Merge cells
dest_sheet.merge_cells('C1:G1')
dest_sheet.merge_cells('H1:L1')
dest_sheet.merge_cells('M1:N1')
dest_sheet.merge_cells('C2:G2')
dest_sheet.merge_cells('H2:L2')
dest_sheet.merge_cells('M2:N2')
dest_sheet.merge_cells('A1:A3')
dest_sheet.merge_cells('B1:B3')
dest_sheet.merge_cells('R1:R3')
dest_sheet.merge_cells('S1:S3')
dest_sheet.merge_cells('T1:T3')
# Center-align all cells in destination sheet
for row in dest_sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# Set the width of column B to 15
dest_sheet.column_dimensions['B'].width = 15
dest_sheet.column_dimensions['O'].width = 12
dest_sheet.column_dimensions['P'].width = 12
dest_sheet.column_dimensions['Q'].width = 12
dest_sheet.column_dimensions['R'].width = 12
dest_sheet.column_dimensions['S'].width = 12

#colors
green_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
blue_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
orange_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

# Apply green fill only to the first two rows for all columns
for row in range(1, 4):  # Rows 1 and 3
    for col in range(1, dest_sheet.max_column + 1):  # All columns
        dest_sheet.cell(row=row, column=col).fill = green_fill

# Apply blue fill to the first two columns from the 4th row to the end
for row in range(4, dest_sheet.max_row + 1):  # Starting from row 4 to the end
    for col in range(1, 3):  # First two columns
        dest_sheet.cell(row=row, column=col).fill = blue_fill

# Apply orange fill to the first two columns from the 4th row to the end
for row in range(4, dest_sheet.max_row + 1):  # Starting from row 4 to the end
    for col in range(3, 21):  # First two columns
        dest_sheet.cell(row=row, column=col).fill = orange_fill

#---------------------------- Student's Report Sheet ---------------------------
# Create new sheet with name "Student's Report"
dest_sheet1 = dest_workbook.create_sheet("Student's Report")
Info_Sheet = wb['Information']


dest_sheet1.cell(row=1, column=1, value="Student's Report")
dest_sheet1.cell(row=3, column=1, value="Name")
dest_sheet1.cell(row=5, column=1, value="Course Title")
dest_sheet1.cell(row=5, column=4, value="Course Number")
dest_sheet1.cell(row=7, column=1, value="Semester")
dest_sheet1.cell(row=7, column=4, value="Year")
dest_sheet1.cell(row=9, column=1, value="Summary")
dest_sheet1.cell(row=11, column=2, value="Total")
dest_sheet1.cell(row=11, column=3, value="Letter Grade")
dest_sheet1.cell(row=11, column=4, value="Pass / Fail")
dest_sheet1.cell(row=13, column=1, value="Details")
dest_sheet1.cell(row=14, column=1, value="Assignments")
dest_sheet1.cell(row=15, column=1, value="A1")
dest_sheet1.cell(row=15, column=2, value="A2")
dest_sheet1.cell(row=15, column=3, value="A3")
dest_sheet1.cell(row=15, column=4, value="A4")
dest_sheet1.cell(row=15, column=5, value="A5")
dest_sheet1.cell(row=17, column=1, value="Quizzes")
dest_sheet1.cell(row=18, column=1, value="Q1")
dest_sheet1.cell(row=18, column=2, value="Q2")
dest_sheet1.cell(row=18, column=3, value="Q3")
dest_sheet1.cell(row=18, column=4, value="Q4")
dest_sheet1.cell(row=18, column=5, value="Q5")
dest_sheet1.cell(row=20, column=1, value="Midterm 1")
dest_sheet1.cell(row=20, column=2, value="Midterm 2")
dest_sheet1.cell(row=20, column=3, value="Presentation")
dest_sheet1.cell(row=20, column=4, value="Project")
dest_sheet1.cell(row=20, column=5, value="final Exam")

# Paste the value into cell in column 3, row 5 of the destination sheet
dest_sheet1.cell(row=5, column=2).value = Info_Sheet.cell(row=2, column=2).value
dest_sheet1.cell(row=5, column=5).value = Info_Sheet.cell(row=3, column=2).value
dest_sheet1.cell(row=7, column=2).value = Info_Sheet.cell(row=4, column=2).value
dest_sheet1.cell(row=7, column=5).value = Info_Sheet.cell(row=5, column=2).value

# Create a list to store names from B4 to B34
name_list = [Full_Sheet.cell(row=i, column=2).value for i in range(4, 35)]

# Create a comma-separated string from the list
name_str = ', '.join(name_list)

# Create DataValidation object
dv = DataValidation(type="list", formula1=f'"{name_str}"', allow_blank=True)

# Apply data validation to cell B3 in Info_Sheet
dest_sheet1.add_data_validation(dv)
dv.add(dest_sheet1['B3'])

dest_sheet1["B12"].value = f'=INDEX(\'Letter Grades\'!R4:R34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["C12"].value = f'=INDEX(\'Letter Grades\'!S4:S34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["A16"].value = f'=INDEX(\'Letter Grades\'!C4:C34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["B16"].value = f'=INDEX(\'Letter Grades\'!D4:D34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["C16"].value = f'=INDEX(\'Letter Grades\'!E4:E34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["D16"].value = f'=INDEX(\'Letter Grades\'!F4:F34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["E16"].value = f'=INDEX(\'Letter Grades\'!G4:G34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["A19"].value = f'=INDEX(\'Letter Grades\'!H4:H34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["B19"].value = f'=INDEX(\'Letter Grades\'!I4:I34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["C19"].value = f'=INDEX(\'Letter Grades\'!J4:J34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["D19"].value = f'=INDEX(\'Letter Grades\'!K4:K34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["E19"].value = f'=INDEX(\'Letter Grades\'!L4:L34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["A21"].value = f'=INDEX(\'Letter Grades\'!M4:M34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["B21"].value = f'=INDEX(\'Letter Grades\'!N4:N34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["C21"].value = f'=INDEX(\'Letter Grades\'!O4:O34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["D21"].value = f'=INDEX(\'Letter Grades\'!P4:P34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["E21"].value = f'=INDEX(\'Letter Grades\'!Q4:Q34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
dest_sheet1["D12"].value = f'=INDEX(\'Letter Grades\'!T4:T34, MATCH(B3, \'Letter Grades\'!B4:B34, 0))'
# Create a bar chart
chart1 = BarChart()

# Define data and categories
data = Reference(dest_sheet1, min_col=1, min_row=16, max_col=5, max_row=16)
categories = Reference(dest_sheet1, min_col=1, min_row=14, max_col=1, max_row=14)


# Add data and categories to the chart
chart1.add_data(data, titles_from_data=False)
chart1.set_categories(categories)

chart1.legend = None

# Position the chart on the worksheet
dest_sheet1.add_chart(chart1, "A23")

# Create a bar chart
chart2 = BarChart()

# Define data and categories
data = Reference(dest_sheet1, min_col=1, min_row=19, max_col=5, max_row=19)
categories = Reference(dest_sheet1, min_col=1, min_row=17, max_col=1, max_row=17)


# Add data and categories to the chart
chart2.add_data(data, titles_from_data=False)
chart2.set_categories(categories)

chart2.legend = None

# Position the chart on the worksheet
dest_sheet1.add_chart(chart2, "A40")


dest_sheet1.merge_cells('A1:E1')
dest_sheet1.merge_cells('A2:E2')
dest_sheet1.merge_cells('A4:E4')
dest_sheet1.merge_cells('B3:E3')
dest_sheet1.merge_cells('A8:E8')
dest_sheet1.merge_cells('A9:E9')
dest_sheet1.merge_cells('A10:E10')
dest_sheet1.merge_cells('A13:E13')
dest_sheet1.merge_cells('A14:E14')
dest_sheet1.merge_cells('A17:E17')

# Set the width of column A to E = 15
dest_sheet1.column_dimensions['A'].width = 14.5
dest_sheet1.column_dimensions['B'].width = 14.5
dest_sheet1.column_dimensions['C'].width = 14.5
dest_sheet1.column_dimensions['D'].width = 14.5
dest_sheet1.column_dimensions['E'].width = 14.5

# Center-align all cells in destination sheet
for row in dest_sheet1.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

dest_sheet1.cell(row=1, column=1).fill = green_fill
dest_sheet1.cell(row=3, column=1).fill = green_fill
dest_sheet1.cell(row=5, column=1).fill = green_fill
dest_sheet1.cell(row=5, column=4).fill = green_fill
dest_sheet1.cell(row=7, column=1).fill = green_fill
dest_sheet1.cell(row=7, column=4).fill = green_fill
dest_sheet1.cell(row=9, column=1).fill = green_fill
dest_sheet1.cell(row=11, column=2).fill = green_fill
dest_sheet1.cell(row=11, column=3).fill = green_fill
dest_sheet1.cell(row=11, column=4).fill = green_fill
dest_sheet1.cell(row=13, column=1).fill = green_fill
dest_sheet1.cell(row=14, column=1).fill = green_fill
dest_sheet1.cell(row=15, column=1).fill = green_fill
dest_sheet1.cell(row=15, column=2).fill = green_fill
dest_sheet1.cell(row=15, column=3).fill = green_fill
dest_sheet1.cell(row=15, column=4).fill = green_fill
dest_sheet1.cell(row=15, column=5).fill = green_fill
dest_sheet1.cell(row=17, column=1).fill = green_fill
dest_sheet1.cell(row=18, column=1).fill = green_fill
dest_sheet1.cell(row=18, column=2).fill = green_fill
dest_sheet1.cell(row=18, column=3).fill = green_fill
dest_sheet1.cell(row=18, column=4).fill = green_fill
dest_sheet1.cell(row=18, column=5).fill = green_fill
dest_sheet1.cell(row=20, column=1).fill = green_fill
dest_sheet1.cell(row=20, column=2).fill = green_fill
dest_sheet1.cell(row=20, column=3).fill = green_fill
dest_sheet1.cell(row=20, column=4).fill = green_fill
dest_sheet1.cell(row=20, column=5).fill = green_fill
dest_sheet1.cell(row=3, column=2).fill = blue_fill



dest_filename = "/Users/ehf/Desktop/essafinal.xlsx"
dest_workbook.save(dest_filename)

wb.close()
dest_workbook.close()
